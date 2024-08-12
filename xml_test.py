import os
from openpyxl import Workbook
import openpyxl
import numpy as np

def get_lems(path):
    'get element block'

    num_ele = []
    count = 0
    text = []

    with open (path, "r") as f:
        for row in f:
            text.append(row.strip())
            count +=1
            if "EBLOCK" in row:
                num = [i.strip() for i in row.split((","))]
                #print(num)
                num_ele = int(num[3])
                number = count+1
    #print(text)
    text2 = text[number:number+(2*num_ele)]
    #print(text2)
    lines_anzahl = list(range(len(text2)))
    el = lines_anzahl[0:len(lines_anzahl):2]
    elems = [text2[i] for i in el]
    #print(elems)

    return elems

def get_mat_value(text):
    'get the Material DATA from Materialblock'
    value_line = []
    value=[]
    with open (text, "r") as r:
        for row in r:
            if "MPDATA,R5.0, 1,EX" in row:
                row.strip()
                value_line.append(list(row.split(",")))
    for i in value_line:
        #print(i)
        value.append(float(i[6].strip()))
                #value_line.append(float((row[30:47].strip())))
    sorted_values = sorted(value,reverse=True)
    return sorted_values

def mat(path):
    #element lines from cdb
    elems = get_lems(path)
    #print(elems)

    #values from material bins
    values = get_mat_value(path)

    #values as string
    #values_str = []
    #for i in values:
        #values_str.append(str(i))

    values_str = [str(i) for i in values]

    #extract material number from string line ele
    mat = []
    #print(elems)
    #print(len(elems))
    for i in elems:
        #print(i)
        #print(i[0:7])
        li = i[0:7].strip()
        lii = li.split()
        if len(lii) > 1:
            li = i[0:4].strip()
        mat.append(int(li))

    #anzahl materialien pro bin
    count = 1
    p = []
    s = sorted(mat)[:]
    s.append(0)
    for i in range(len(s) - 1):
        if s[i] == s[i + 1]:
            count += 1
        elif s[i] != s[i + 1]:
            p.append(count)
            count = 1

    #list with grouped values for each element
    new_moduli = []
    for i in range(len(values_str)):  # 'i = 0,1'
        for j in range(p[i]):  #
            new_moduli.append(values_str[i])


    #Anzahl MAterialien pro bin, Bins, Material für jedes Element, Anzahl Elemente, new_moduli
    return p, len(p), sorted(mat),len(sorted(mat,reverse=True)), values, new_moduli

def volume():

    #get volume data from text files
    with open("V.txt", "r") as v:
        text = v.read()

    s = text.split('\n')
    ss = [i.split(":") for i in s]

    text = [i[0] for i in ss]
    v = [i[1] for i in ss]

    return text,v

def csv(path,group):

    #create workbook class
    wb = Workbook()
    #activate sheet
    ws = wb.active

    #get elems ad material from cbd
    elems = mat(path[0])
    values = elems[4]
    len_elems = elems[3]
    len_bins = elems[1]
    elems_pro_bin = elems[0]
    new_moduli = [float(i) for i in elems[5]]

    #get data volume
    text,v = volume()

    #get R
    with open("R.txt", "r") as r:
        r = r.read()
    R = float(r)

    #fill sheet with data old version (Bonemat)
    ws["A1"] ="max"
    ws["A2"] ="min"
    ws["A3"] ="mean"

    ws["A5"] = "elements pro bin"
    ws["B5"] = "Value Bin Old"
    ws["C5"] = "Value grouped Element MPa"

    ws["B1"] = max(values)
    ws["B2"] = min(values)
    ws["B3"] = sum(values)/len(values)

    ws["C1"] = max(new_moduli)
    ws["C2"] = min(new_moduli)
    ws["C3"] = sum(new_moduli)/len(new_moduli)

    for j in range(len(elems_pro_bin)):
        ws["A"+str(j+6)] = elems_pro_bin[j]
    for k in range(len(values)):
        ws["B"+str(k+6)]=values[k]
    for l in range(len(new_moduli)):
        ws["C"+str(l+6)]=new_moduli[l]

    #fill sheet with data from new version
    elems2 = mat(path[1])
    values2 = elems2[4]
    len_elems = elems2[3]
    len_bins = elems2[1]
    elems_pro_bin2 = elems2[0]
    new_moduli2 = [float(i) for i in elems2[5]]

    ws["F1"] = "max"
    ws["F2"] = "min"
    ws["F3"] = "mean"

    ws["F5"] = "elements pro bin"
    ws["G5"] = "Value Bin New"
    ws["H5"] = "Value grouped Element MPa"

    ws["G1"] = max(values2)
    ws["G2"] = min(values2)
    ws["G3"] = sum(values2) / len(values2)

    ws["H1"] = max(new_moduli2)
    ws["H2"] = min(new_moduli2)
    ws["H3"] = sum(new_moduli2) / len(new_moduli2)

    for j in range(len(elems_pro_bin2)):
        ws["F" + str(j + 6)] = elems_pro_bin2[j]
    for k in range(len(values2)):
        ws["G" + str(k + 6)] = values2[k]
    for l in range(len(new_moduli2)):
        ws["H" + str(l + 6)] = new_moduli2[l]

    #fill shet with datat from calculation (Volume und R- none bone Fraktion)
    ws["J1"] = str(text[0])
    ws["J2"] = str(text[1])
    ws["J3"] = str(text[2])
    ws['J4'] = "v under HU 50"

    ws["K1"] = round(float(v[0]),4)
    ws["K2"] = round(float(v[1]),4)
    ws["K3"] = round(float(v[2]),4)
    ws["K4"] = str(round(float((100*float(v[0]))/float(v[2])),4))

    ws["M1"] = "R:"
    ws["N1"] = str(round(float(R),4))

    #data for R and HU/ E, for plot the charts
    hu = list(range(51))
    Q = [-0.003935729 + 0.000791701 * i for i in hu]

    #Non bone fraction R
    R_1 = np.arange(0.01, 0.2, 0.005)
    # for HU 50 = Q : 0,035
    E_RR = np.array([(10 ** (-8.54 + (4.05 * (np.log10(400 * (0.035 / (0.035 + (1.4 * s))))))) * 1000) for s in R_1])
    E_R = E_RR[np.isfinite(E_RR)]

    #new E-Mod and del NAN (because log equation)
    EE = np.array([(10 ** (-8.54 + (4.05 * (np.log10(400 * (q / (q + 1.4 * R)))))) * 1000) for q in Q])
    E = EE[np.isfinite(EE)]
    # e = tuple(E)

    ##orig E-Modul,
    E_orig = np.array([(6850 * ((((0.87719 * r) + 0.07895) / 0.6) ** 1.49)) for r in Q])

    #first chart with R-E correlation
    chart3 = openpyxl.chart.LineChart()
    chart3.style = 10
    chart3.x_axis.title = "R"
    chart3.y_axis.title = "E in MPa"
    #chart3.x_axis.scaling.min = 0
    #chart3.x_axis.scaling.max = 0.2
    chart3.title = None

    ws['AE5'] = "E-R"
    ws['AF5'] = "R"
    #fill rows with data
    for row in range(len(E_R)):
        ws.cell(row = 6+int(row), column=31).value = E_R[row]
    for row in range(len(R_1)):
        ws.cell(row = 6+int(row), column=32).value = R_1[row]

    #get data from rows
    x = openpyxl.chart.Reference(ws, min_col=32, max_col = 32, min_row=6, max_row=len(R_1)+5)
    y = openpyxl.chart.Reference(ws, min_col=31, max_col = 31, min_row=5, max_row=len(E_R)+5)
    chart3.add_data(y, titles_from_data=True)
    #set new x values (R)
    chart3.set_categories(x)
    ws.add_chart(chart3, 'T22')

    #chart for Group-E.mod
    chart0 = openpyxl.chart.LineChart()
    chart0.style = 10
    chart0.x_axis.title = "HU"
    chart0.y_axis.title = "E in MPa"

    # add datat to chart
    ws['AC5'] = "E-new"
    for row in range(len(E)):
        ws.cell(row = 6+int(row), column = 29).value = E[row]
    values0 = openpyxl.chart.Reference(ws, min_col=29, max_col=29, min_row=5, max_row=len(E)+5)
    chart0.add_data(values0, titles_from_data=True)

    ws['AD5'] = "E-orig"
    for row in range(len(E_orig)):
        ws.cell(row=6 + int(row), column=30).value = E_orig[row]
    values00 = openpyxl.chart.Reference(ws, min_col=30, max_col=30, min_row=5, max_row=len(E_orig)+5)
    chart0.add_data(values00, titles_from_data=True)

    # add chart to worksheet
    ws.add_chart(chart0, "T7")


    #chart for Bin-E-mod
    chart = openpyxl.chart.LineChart()
    chart.style = 10
    chart.x_axis.title = "Gruppierung"
    chart.y_axis.title = "E [MPa]"
    #data for chart
    values1 = openpyxl.chart.Reference(ws, min_col=2, max_col=2, min_row=5, max_row=len(values)+5)
    values2 = openpyxl.chart.Reference(ws, min_col=7, max_col=7, min_row=5, max_row=len(values)+5)
    #add datat to chart
    chart.add_data(values1, titles_from_data=True)
    chart.add_data(values2, titles_from_data=True)
    #add chart to worksheet
    ws.add_chart(chart, "K7")


    chart2 = openpyxl.chart.LineChart()
    chart2.style = 10
    chart2.x_axis.title = "Gruppierung"
    chart2.y_axis.title = "Anzahl Elemente"

    values11 = openpyxl.chart.Reference(ws, min_col=1, max_col=1, min_row=5, max_row=len(values) + 5)
    values22 = openpyxl.chart.Reference(ws, min_col=6, max_col=6, min_row=5, max_row=len(values) + 5)
    chart2.add_data(values11,titles_from_data=True)
    chart2.add_data(values22,titles_from_data=True)
    ws.add_chart(chart2, "K22")


    #path for save
    path_ord = r"C:/Users/8iries/Downloads/test_cdb_trash"
    os.chdir(path_ord)
    name = path[0].split("\\")
    name2 = name[-1]

    #save excel data
    wb.save(name2+".xlsx")

"""
#TEXT DATA with max_mean_min 

    if group == 0:
        with open ("BM_moduli.txt","w") as w:
            w.write("max :" + str(max(new_moduli)))
            w.write("\nmin :" + str(min(new_moduli)))
            w.write("\nmean :" + str(sum(new_moduli) / len(new_moduli)))
            w.write("\nlen :" + str(len(new_moduli)))
            for l in new_moduli:
                w.write("\n")
                w.write(str(round(l)))

    elif group == 1:

        with open ("BM_set_moduli.txt","w") as w:
            w.write("max :" + str(max(values)))
            w.write("\nmin :" + str(min(values)))
            w.write("\nmean :" + str(sum(values) / len(values)))
            w.write("\nlen :" + str(len(values)))
            for l in values:
                w.write("\n")
                w.write(str(round(l)))

        with open ("BM_new_moduli.txt","w") as q:
            q.write("max :" + str(max(new_moduli)))
            q.write("\nmin :" + str(min(new_moduli)))
            q.write("\nmean :" + str(sum(new_moduli) / len(new_moduli)))
            q.write("\nlen :" + str(len(new_moduli)))
            for l in new_moduli:
                q.write("\n")
                q.write(str(round(l)))

        with open ("BM_len_elems_pro_bin.txt","w") as p:
            for l in elems_pro_bin:
                p.write(str(l))
                p.write("\n")


"""

#1 = original, 2 = new
path_1 = [r"C:\Users\8iries\Desktop\Ba\paper\1\mapped\sacrumv2\sacrumv2_p1.cdb",r"C:\Users\8iries\Desktop\Ba\paper\1\mapped\sacrumv2\sacrumv2_p1_HU_sep_one.cdb"]

#path = r"C:\Users\8iries\Downloads\test_cdb_trash\111.cdb"

#0 für Gap = 0, 1 für GAp
csv(path_1,1)
